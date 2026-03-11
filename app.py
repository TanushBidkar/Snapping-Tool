from flask import Flask, render_template, request, jsonify, session
import pandas as pd
import os
import json


app = Flask(__name__)
app.secret_key = 'boq_tool_secret_key'

UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'})

    file = request.files['file']

    if file.filename == '':
        return jsonify({'success': False, 'error': 'No file selected'})

    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({'success': False, 'error': 'Please upload an Excel file (.xlsx or .xls)'})

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    file.save(filepath)
    firebase_url = None

    # Read sheet names from the Excel file
    try:
        xl = pd.ExcelFile(filepath)
        sheet_names = xl.sheet_names

        # Check if BOQ and Measurement Sheet exist
        boq_exists = any('BOQ' in s for s in sheet_names)
        measurement_exists = any('Measurement' in s for s in sheet_names)

        return jsonify({
    'success': True,
    'filename': file.filename,
    'sheets': sheet_names,
    'boq_exists': boq_exists,
    'measurement_exists': measurement_exists,
    'firebase_url': firebase_url
})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/get-boq', methods=['POST'])
def get_boq():
    data = request.json
    filename = data.get('filename')
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

    try:
        xl = pd.ExcelFile(filepath)
        # Find the BOQ sheet
        print("Available sheets:", xl.sheet_names)
        boq_sheet = next((s for s in xl.sheet_names if 'BOQ' in s.upper()), None)
        print("BOQ sheet found:", boq_sheet)

        if not boq_sheet:
            return jsonify({'success': False, 'error': 'BOQ sheet not found'})

        df = pd.read_excel(filepath, sheet_name=boq_sheet, header=None)
        df = df.fillna('')

        rows = df.values.tolist()
        # Convert all values to string for JSON
        rows = [[str(cell) for cell in row] for row in rows]

        return jsonify({'success': True, 'data': rows, 'sheet_name': boq_sheet})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/get-measurement', methods=['POST'])
def get_measurement():
    data = request.json
    filename = data.get('filename')
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

    try:
        xl = pd.ExcelFile(filepath)
        # Find the Measurement sheet
        measurement_sheet = next((s for s in xl.sheet_names if 'measurement' in s.lower()), None)
        print("Measurement sheet found:", measurement_sheet)

        if not measurement_sheet:
            return jsonify({'success': False, 'error': 'Measurement Sheet not found'})

        df = pd.read_excel(filepath, sheet_name=measurement_sheet, header=None)
        df = df.fillna('')

        rows = df.values.tolist()
        rows = [[str(cell) for cell in row] for row in rows]

        return jsonify({'success': True, 'data': rows, 'sheet_name': measurement_sheet})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})
    
@app.route('/upload-photo', methods=['POST'])
def upload_photo():
    if 'photo' not in request.files:
        return jsonify({'success': False, 'error': 'No photo uploaded'})

    photo = request.files['photo']
    item_name = request.form.get('item_name', 'unknown')
    excel_filename = request.form.get('excel_filename', 'unknown')

    # Clean item name for use as filename
    import re
    from datetime import datetime
    clean_item = re.sub(r'[^a-zA-Z0-9_\-]', '_', item_name)[:50]
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    photo_filename = f"{timestamp}_{clean_item}{os.path.splitext(photo.filename)[1]}"

    photo_path = os.path.join('uploads', photo_filename)
    photo.save(photo_path)
    return jsonify({'success': True, 'url': None})
@app.route('/save-measurement', methods=['POST'])
def save_measurement():
    data = request.json
    filename = data.get('filename')
    item_name = data.get('item_name', '')
    # Expect decimal feet values from frontend (e.g. 0.38)
    length_ft = data.get('length_ft', 0)
    breadth_ft = data.get('breadth_ft', 0)
    area_sqft = data.get('area_sqft', 0)
    # Optional photographic fields
    photo_nos = data.get('photo_nos')
    photo_L = data.get('photo_L')
    photo_B = data.get('photo_B')
    photo_DH = data.get('photo_DH')

    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)

    try:
        xl = pd.ExcelFile(filepath)
        measurement_sheet = next((s for s in xl.sheet_names if 'measurement' in s.lower()), None)

        df = pd.read_excel(filepath, sheet_name=measurement_sheet, header=None)

        # Find the row matching item_name and update Photographic Dimension and QTY columns
        item_parts = [p.strip() for p in item_name.split('›') if p.strip()]
        search_term = item_parts[0] if item_parts else item_name

        # Locate Photographic Dimension section: find header and columns within that section
        col_idxs = {'Nos': None, 'L': None, 'B': None, 'D/H': None, 'QTY': None}

        print(f"DEBUG col_idxs: {col_idxs}")
        print(f"DEBUG search_term: {search_term}")
        print(f"DEBUG photo_L={photo_L}, photo_B={photo_B}, photo_DH={photo_DH}, area={area_sqft}")
        
        # First, find the "Photographic Dimension" header cell location
        photo_dim_col_start = None
        for col in range(df.shape[1]):
            for r in range(min(10, df.shape[0])):
                v = str(df.iat[r, col]).strip().lower()
                if 'photographic' in v and 'dimension' in v:
                    photo_dim_col_start = col
                    break
            if photo_dim_col_start is not None:
                break
        
        # If Photographic Dimension header found, search for columns AFTER it
        # Otherwise search from right to left for the rightmost set of columns
        if photo_dim_col_start is not None:
            search_range = range(photo_dim_col_start, df.shape[1])
        else:
            search_range = range(df.shape[1] - 1, -1, -1)  # right to left
        
        for col in search_range:
            for r in range(min(10, df.shape[0])):
                v = str(df.iat[r, col]).strip().lower()
                if v in ('nos', 'no', 'nos.') and col_idxs['Nos'] is None:
                    col_idxs['Nos'] = col
                elif v == 'l' and col_idxs['L'] is None:
                    col_idxs['L'] = col
                elif v == 'b' and col_idxs['B'] is None:
                    col_idxs['B'] = col
                elif v in ('d/h', 'd h', 'd h.') and col_idxs['D/H'] is None:
                    col_idxs['D/H'] = col
                elif v in ('qty', 'q t y', 'q.ty', 'quantity'):
                    col_idxs['QTY'] = col  # keep overwriting → gets rightmost QTY

        if col_idxs['QTY'] is None:
            col_idxs['QTY'] = df.shape[1] - 1

        # Use exact row index from frontend if provided
        row_index_from_frontend = data.get('row_index', None)

        updated = False
        for idx, row in df.iterrows():
            if row_index_from_frontend is not None:
                if idx != int(row_index_from_frontend):
                    continue
            else:
                row_text = ' '.join([str(c) for c in row.values])
                if search_term.lower() not in row_text.lower():
                    continue

            if col_idxs['Nos'] is not None and photo_nos not in (None, ''):
                df.iat[idx, col_idxs['Nos']] = photo_nos
            if col_idxs['L'] is not None:
                df.iloc[:, col_idxs['L']] = df.iloc[:, col_idxs['L']].astype(object)
                df.iat[idx, col_idxs['L']] = 'NA' if photo_L == 'NA' else round(float(length_ft), 3)
            if col_idxs['B'] is not None:
                df.iloc[:, col_idxs['B']] = df.iloc[:, col_idxs['B']].astype(object)
                df.iat[idx, col_idxs['B']] = 'NA' if photo_B == 'NA' else round(float(breadth_ft), 3)
            if col_idxs['D/H'] is not None:
                df.iloc[:, col_idxs['D/H']] = df.iloc[:, col_idxs['D/H']].astype(object)
                df.iat[idx, col_idxs['D/H']] = photo_DH if photo_DH not in (None, '', 'NA') else 'NA'
            if col_idxs['QTY'] is not None:
                df.iloc[:, col_idxs['QTY']] = df.iloc[:, col_idxs['QTY']].astype(object)
                from openpyxl.utils import get_column_letter
                qty_col = col_idxs['QTY']
                excel_row = idx + 1  # 1-indexed Excel row
                nos_col_letter = get_column_letter(col_idxs['Nos'] + 1) if col_idxs['Nos'] is not None else None
                l_col_letter   = get_column_letter(col_idxs['L'] + 1)   if col_idxs['L']   is not None else None
                b_col_letter   = get_column_letter(col_idxs['B'] + 1)   if col_idxs['B']   is not None else None
                dh_col_letter  = get_column_letter(col_idxs['D/H'] + 1) if col_idxs['D/H'] is not None else None

                has_L  = photo_L  not in (None, '', 'NA')
                has_B  = photo_B  not in (None, '', 'NA')
                has_DH = photo_DH not in (None, '', 'NA')

                if has_L and has_B and nos_col_letter:
                    formula = f"={nos_col_letter}{excel_row}*{l_col_letter}{excel_row}*{b_col_letter}{excel_row}"
                elif has_L and has_DH and nos_col_letter:
                    formula = f"={nos_col_letter}{excel_row}*{l_col_letter}{excel_row}*{dh_col_letter}{excel_row}"
                elif has_B and has_DH and nos_col_letter:
                    formula = f"={nos_col_letter}{excel_row}*{b_col_letter}{excel_row}*{dh_col_letter}{excel_row}"
                elif has_L and nos_col_letter:
                    formula = f"={nos_col_letter}{excel_row}*{l_col_letter}{excel_row}"
                elif has_B and nos_col_letter:
                    formula = f"={nos_col_letter}{excel_row}*{b_col_letter}{excel_row}"
                else:
                    formula = round(float(area_sqft), 3)

                df.iat[idx, qty_col] = formula

            updated = True
            break

        # If not found, append a new row
        if not updated:
            new_row = [''] * df.shape[1]
            new_row[0] = item_name
            if col_idxs['Nos'] is not None and photo_nos not in (None, ''):
                new_row[col_idxs['Nos']] = photo_nos
            if col_idxs['L'] is not None:
                df.iat[idx, col_idxs['L']] = 'NA' if photo_L == 'NA' else str(round(float(length_ft), 3))
            if col_idxs['B'] is not None:
                df.iat[idx, col_idxs['B']] = 'NA' if photo_B == 'NA' else str(round(float(breadth_ft), 3))
                new_row[col_idxs['D/H']] = photo_DH if photo_DH not in (None, '') else 'NA'
            # QTY always = area
            new_row[col_idxs['QTY']] = round(float(area_sqft), 2)
            df.loc[df.shape[0]] = new_row

        # Save updated Excel
        # Save updated Excel — preserve all formatting, only update cell values
        # Save updated Excel — preserve formatting, handle merged cells safely
        from openpyxl import load_workbook
        from openpyxl.cell.cell import MergedCell
        wb = load_workbook(filepath)
        ws = wb[measurement_sheet]

        target_cols = {v for v in col_idxs.values() if v is not None}
        print(f"DEBUG final col_idxs: {col_idxs}")
        print(f"DEBUG target_cols (0-indexed): {target_cols}")

        # Only write to the single target row, not all rows
        target_row_idx = int(row_index_from_frontend) if row_index_from_frontend is not None else None
        if target_row_idx is not None:
            for c_idx in target_cols:
                value = df.iat[target_row_idx, c_idx]
                cell = ws.cell(row=target_row_idx + 1, column=c_idx + 1)
                if isinstance(cell, MergedCell):
                    continue
                if value == '' or value is None:
                    pass
                elif isinstance(value, str) and value == 'NA':
                    cell.value = 'NA'
                elif isinstance(value, str) and value.startswith('='):
                    cell.value = value
                elif isinstance(value, (int, float)) and not (isinstance(value, float) and pd.isna(value)):
                    try:
                        cell.value = float(value)
                    except (TypeError, ValueError):
                        cell. Value = str(value)
        wb.save(filepath)

        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/download-excel')
def download_excel():
    from flask import send_file
    filename = request.args.get('filename')
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    return send_file(filepath, as_attachment=True, download_name=filename)

if __name__ == '__main__':
    app.run(debug=True)
